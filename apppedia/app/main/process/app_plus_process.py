from urllib.parse import quote_plus
from urllib.request import urlopen
from bs4 import BeautifulSoup
from selenium import webdriver

import openpyxl
import time

from .. import db
from app.main import db
from app.main.model.application import Application
from ..service.application_service import application_save

def app_plus_process():
    ''' Application Plus Processing '''
    application_file = openpyxl.load_workbook("/home/apppedia/apppedia/application_list.xlsx")
    application_sheet = application_file['Sheet']
    row = 1
    col = application_sheet['A']

    for cell in col:
        try:
            print(str(application_sheet.cell(row,1).value))
            application = Application.query.filter_by(public_id=str(application_sheet.cell(row,1).value)).first()
            if application:
                html = urlopen('https://search.naver.com/search.naver?where=post&sm=tab_jum&query=' + quote_plus(str(application.developer_name) + ' ' + str(application.name) + ' 장단점 리뷰'))
                soup = BeautifulSoup(html, 'html.parser')
                title = soup.find_all(class_='sh_blog_title')
                for i in title:
                    application.related_name = i.attrs['title']
                    application.related_link = i.attrs['href']
                    break
                print("Application Plus Processing : Modify")
                db.session.commit()
            row += 1
            time.sleep(1.0)
        except:
            print("Application Plus Processing : Fail")
            row += 1
            time.sleep(10.0)
            continue