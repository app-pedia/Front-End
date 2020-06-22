from urllib.request import urlopen
from bs4 import BeautifulSoup

import openpyxl
import requests
import json
import time

from .. import db
from app.main import db
from app.main.model.application import Application
from ..service.application_service import application_save

def app_info_process():
    ''' Application Info Processing '''
    application_file = openpyxl.load_workbook("/home/apppedia/apppedia/application_list.xlsx")
    application_sheet = application_file['Sheet']
    row = 1
    col = application_sheet['A']

    for cell in col:
        try:
            print(str(application_sheet.cell(row,1).value)) 
            data = {'public_id': 'null', 'name': 'null', 'category': 'null', 'developer_name': 'null', 'developer_public_id': 'null', 'rating_total': 'null', 'rating_average': 'null', 'rating_five': 'null', 'rating_four': 'null', 'rating_three': 'null', 'rating_two': 'null', 'rating_one': 'null', 'install': 'null', 'install_link': 'null', 'image_logo': 'null', 'price': 'null', 'update_date': 'null', 'size': 'null', 'version_current': 'null', 'version_needs': 'null', 'contents_grade': 'null', 'interaction': 'null', 'in_app_products': 'null', 'related_name': 'null', 'related_link': 'null'}

            s = requests.session()
            base_url = 'https://www.androidrank.org/api/application/'+str(application_sheet.cell(row,1).value)+'?key=PUT_YOUR_ANDROIDRANK_API_KEY_HERE'
            con = s.get(base_url)
            json_data = json.loads(con.text)

            html = urlopen('https://play.google.com/store/apps/details?id='+str(application_sheet.cell(row,1).value)+'&hl=ko')
            bsObject = BeautifulSoup(html, "html.parser",from_encoding='utf-8')
            data['public_id'] = str(application_sheet.cell(row,1).value)
            length = 0

            application = Application.query.filter_by(public_id=data['public_id']).first()
            if not application:
                data['category'] = json_data['category']
                data['developer_name'] = json_data['developer']['name']
                data['developer_public_id'] = json_data['developer']['developer_id']
                data['rating_total'] = json_data['ratings']['current']['count']
                data['rating_five'] = json_data['ratings']['current']['rank5']
                data['rating_four'] = json_data['ratings']['current']['rank4']
                data['rating_three'] = json_data['ratings']['current']['rank3']
                data['rating_two'] = json_data['ratings']['current']['rank2']
                data['rating_one'] = json_data['ratings']['current']['rank1']
                data['install'] = json_data['installs']['current']['count']
                data['install_link'] = 'https://play.google.com/store/apps/details?id=' + json_data['id']
                data['image_logo'] = json_data['images']['logo']
                data['price'] = json_data['price']['current']
                paragraph_data = bsObject.find('h1', class_='AHFaub')
                data['name'] = paragraph_data.text
                paragraph_data = bsObject.find('div', class_='BHMmbe')
                data['rating_average'] = paragraph_data.text
                paragraph_data = bsObject.find('div', class_='IxB2fe')
                processed_data = paragraph_data.find_all('div', class_='hAyfc')
                while length < len(processed_data):
                    processing_data = processed_data[length]
                    process_data = processing_data.find('div', class_='BgcNfc')
                    if process_data.text == '업데이트 날짜':
                        process_data = processing_data.find('span', class_='htlgb')
                        data['update_date'] = process_data.text
                    elif process_data.text == '크기':
                        process_data = processing_data.find('span', class_='htlgb')
                        data['size'] = process_data.text
                    elif process_data.text == '현재 버전':
                        process_data = processing_data.find('span', class_='htlgb')
                        data['version_current'] = process_data.text
                    elif process_data.text == '필요한 Android 버전':
                        process_data = processing_data.find('span', class_='htlgb')
                        data['version_needs'] = process_data.text
                    elif process_data.text == '콘텐츠 등급':
                        process_data = processing_data.find('span', class_='htlgb')
                        process_data = process_data.find_all('div')
                        process_data = process_data[1]
                        data['contents_grade'] = process_data.text
                    elif process_data.text == '상호작용 요소':
                        process_data = processing_data.find('span', class_='htlgb')
                        data['interaction'] = process_data.text
                    elif process_data.text == '인앱 상품':
                        process_data = processing_data.find('span', class_='htlgb')
                        data['in_app_products'] = process_data.text
                    length += 1
                application_save(data=data)
                print("Application Info Processing : Save")
            else:
                application.category = json_data['category']
                application.developer_name = json_data['developer']['name']
                application.developer_public_id = json_data['developer']['developer_id']
                application.rating_total = json_data['ratings']['current']['count']
                application.rating_five = json_data['ratings']['current']['rank5']
                application.rating_four = json_data['ratings']['current']['rank4']
                application.rating_three = json_data['ratings']['current']['rank3']
                application.rating_two = json_data['ratings']['current']['rank2']
                application.rating_one = json_data['ratings']['current']['rank1']
                application.install = json_data['installs']['current']['count']
                application.install_link = 'https://play.google.com/store/apps/details?id=' + json_data['id']
                application.image_logo = json_data['images']['logo']
                application.price = json_data['price']['current']
                paragraph_data = bsObject.find('h1', class_='AHFaub')
                application.name = paragraph_data.text
                paragraph_data = bsObject.find('div', class_='BHMmbe')
                application.rating_average = paragraph_data.text
                paragraph_data = bsObject.find('div', class_='IxB2fe')
                processed_data = paragraph_data.find_all('div', class_='hAyfc')
                while length < len(processed_data):
                    processing_data = processed_data[length]
                    process_data = processing_data.find('div', class_='BgcNfc')
                    if process_data.text == '업데이트 날짜':
                        process_data = processing_data.find('span', class_='htlgb')
                        application.update_date = process_data.text
                    elif process_data.text == '크기':
                        process_data = processing_data.find('span', class_='htlgb')
                        application.size = process_data.text
                    elif process_data.text == '현재 버전':
                        process_data = processing_data.find('span', class_='htlgb')
                        application.version_current = process_data.text
                    elif process_data.text == '필요한 Android 버전':
                        process_data = processing_data.find('span', class_='htlgb')
                        application.version_needs = process_data.text
                    elif process_data.text == '콘텐츠 등급':
                        process_data = processing_data.find('span', class_='htlgb')
                        process_data = process_data.find_all('div')
                        process_data = process_data[1]
                        application.contents_grade = process_data.text
                    elif process_data.text == '상호작용 요소':
                        process_data = processing_data.find('span', class_='htlgb')
                        application.interaction = process_data.text
                    elif process_data.text == '인앱 상품':
                        process_data = processing_data.find('span', class_='htlgb')
                        application.in_app_products = process_data.text
                    length += 1
                db.session.commit()
                print("Application Info Processing : Modify")

            row +=1
            time.sleep(5.0)
        except:
            print("Application Info Processing : Fail")
            row += 1
            time.sleep(50.0)
            continue
