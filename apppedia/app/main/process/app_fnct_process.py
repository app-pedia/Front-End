from urllib.request import urlopen
from bs4 import BeautifulSoup

import openpyxl
import requests
import json
import time
import sys
import logging

from konlpy.tag import Okt
from collections import Counter

from .. import db
from app.main import db
from app.main.model.application import Application
from app.main.model.function import Function
from ..service.function_service import function_save

def app_fnct_process():
    ''' Application Fnct Processing '''
    application_file = openpyxl.load_workbook("/home/apppedia/apppedia/application_list.xlsx")
    application_sheet = application_file['Sheet']
    row = 1
    col = application_sheet['A']
    f_format = ['예술', '디자인', '자동차', '뷰티', '패션', '옷', '책', '비즈니스', '만화', '비즈니스', '커뮤니케이션', '교육', '엔터테인먼트',
                '엔터테인먼트', '경제', '금융', '경영', '음식', '헬스', '건강', '운동', '스포츠', '피트니스', '도서', '독서',
                '인테리어', '라이프스타일', '라이프', '맵', '지도', 'GPS', '네비게이션', '음악', '노래', '의료', '뉴스', '매거진', '잡지',
                '포토', '사진', '동영상', '영상', '채널', '위치', '생산', '쇼핑', '소셜', '미디어', '여행', '비디오', '게임',
                '날씨','미용','음성인식','타투','코딩','메모','메모장','필터','카메라','계산기','시계','시간표','배달','음식','영화','카페',
                '중고','쇼핑몰','코로나','딜리버리','알바','교통','대중교통','문자','메세지','캘린더','켈린','메일','티비','TV']

    for cell in col:
        try:
            print(str(application_sheet.cell(row,1).value))
            data = {'public_id': 'null', 'application_public_id': 'null', 'detail': 'null'}

            html = urlopen('https://play.google.com/store/apps/details?id='+str(application_sheet.cell(row,1).value)+'&hl=ko')
            bsObject = BeautifulSoup(html, "html.parser")
            paragraph_data = bsObject.find('div', jsname='sngebd')
            description = paragraph_data.text
            okt=Okt()
            data['application_public_id'] = str(application_sheet.cell(row,1).value)

            application = Application.query.filter_by(public_id=data['application_public_id']).first()
            if application:
                for format in f_format:
                    if format in description:
                        fnct = okt.nouns(format)
                        data['detail'] = str(fnct)[2:-2]
                        print(data['detail'])
                        function = Function.query.filter_by(application_public_id=data['application_public_id']).filter_by(detail=data['detail']).first()
                        if not function:
                            function_save(data=data)
                            print("Application Fnct Processing : Save")
            
            row +=1
            time.sleep(1.0)
        except:
            print("Application Fnct Processing : Fail")
            row +=1
            time.sleep(10.0)