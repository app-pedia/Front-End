from urllib.request import urlopen
from bs4 import BeautifulSoup

import openpyxl
import requests
import json
import time

from .. import db
from app.main import db
from app.main.model.developer import Developer
from ..service.developer_service import developer_save

def dev_info_app_list_process():
    ''' Developer Info Processing and Application List Processing '''
    developer_file = openpyxl.load_workbook("/home/apppedia/apppedia/developer_list.xlsx")
    developer_sheet = developer_file['Sheet']
    row = 1
    col = developer_sheet['A']
    #application_file = openpyxl.Workbook()
    #application_sheet = application_file.active

    for cell in col:
        try:
            print(str(developer_sheet.cell(row,1).value)) 
            data = {'public_id': 'null', 'name': 'null', 'country': 'null', 'address': 'null', 'web': 'null', 'rating_total': 'null', 'rating_average': 'null', 'install_achieved': 'null'}
            html = urlopen('https://www.androidrank.org/developer?id='+str(developer_sheet.cell(row,1).value))
            bsObject = BeautifulSoup(html, "html.parser")
            paragraph_data = bsObject.find_all('table', class_='appstat')
            developer_data = paragraph_data[0]
            processed_data = developer_data.find_all('tr')
            length = 0
            data['public_id'] = str(developer_sheet.cell(row,1).value)

            developer = Developer.query.filter_by(public_id=data['public_id']).first()
            if not developer:
                while length < len(processed_data):
                    processing_data = processed_data[length]
                    process_data = processing_data.find('th')
                    if process_data.text == 'Title:':
                        process_data = processing_data.find('td')
                        data['name'] = process_data.text
                    elif process_data.text == 'Country:':
                        process_data = processing_data.find('td')
                        data['country'] = process_data.find("a").text
                    elif process_data.text == 'Address:':
                        process_data = processing_data.find('td')
                        data['address'] = process_data.text
                    elif process_data.text == 'Web:':
                        process_data = processing_data.find('td')
                        data['web'] = process_data.find("a")["href"]
                    elif process_data.text == 'Total ratings:':
                        process_data = processing_data.find('td')
                        data['rating_total'] = process_data.text
                    elif process_data.text == 'Average rating:':
                        process_data = processing_data.find('td')
                        data['rating_average'] = process_data.text
                    elif process_data.text == 'Installs (achieved):':
                        process_data = processing_data.find('td')
                        data['install_achieved'] = process_data.text
                    length += 1
                developer_save(data=data)
                print("Developer Info Processing : Save")
            else:
                while length < len(processed_data):
                    processing_data = processed_data[length]
                    process_data = processing_data.find('th')
                    if process_data.text == 'Title:':
                        process_data = processing_data.find('td')
                        developer.name = process_data.text
                    elif process_data.text == 'Country:':
                        process_data = processing_data.find('td')
                        developer.country = process_data.find("a").text
                    elif process_data.text == 'Address:':
                        process_data = processing_data.find('td')
                        developer.address = process_data.text
                    elif process_data.text == 'Web:':
                        process_data = processing_data.find('td')
                        developer.web = process_data.find("a")["href"]
                    elif process_data.text == 'Total ratings:':
                        process_data = processing_data.find('td')
                        developer.rating_total = process_data.text
                    elif process_data.text == 'Average rating:':
                        process_data = processing_data.find('td')
                        developer.rating_average = process_data.text
                    elif process_data.text == 'Installs (achieved):':
                        process_data = processing_data.find('td')
                        developer.install_achieved = process_data.text
                    length += 1
                db.session.commit()
                print("Developer Info Processing : Modify")

            #paragraph_data = bsObject.find_all('td', style='text-align:left;')
            #for a in paragraph_data:
            #    public_id = (a.find("a")["href"])
            #    public_id = public_id.split('/')
            #    application_sheet.append([public_id[3]])
            #    application_file.save('application_list.xlsx')
            #    application_file.close()
            #print("Application List Processing : Finish")

            row += 1
            time.sleep(5.0)
        except:
            print("Developer Info Processing and Application List Processing : Fail")
            row += 1
            time.sleep(50.0)
            continue